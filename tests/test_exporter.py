from __future__ import annotations

import csv
import unittest
from pathlib import Path
from tempfile import TemporaryDirectory

from app.config import Settings
from app.services.exporter import Exporter


class FakeDb:
    def __init__(self, findings: list[dict]):
        self.findings = findings

    def list_findings(self, run_id: int | None = None, limit: int = 500) -> list[dict]:
        return self.findings

    def list_findings_for_cases(self, case_ids: list[int]) -> list[dict]:
        return self.findings


class ExporterTests(unittest.TestCase):
    def test_csv_export_is_user_facing(self) -> None:
        with TemporaryDirectory() as temp_dir:
            exporter = Exporter(
                Settings(export_dir=Path(temp_dir)),
                FakeDb([self._finding()]),  # type: ignore[arg-type]
            )

            path = exporter.csv_for_run(1)

            with path.open("r", encoding="utf-8-sig", newline="") as handle:
                row = next(csv.DictReader(handle))

        self.assertIn("Домен", row)
        self.assertIn("Почему отмечен", row)
        self.assertNotIn("ID дела", row)
        self.assertNotIn("SHA-256 HTML", row)
        self.assertEqual(row["Категория"], "Фишинг")

    def test_xlsx_export_keeps_technical_evidence_on_second_sheet(self) -> None:
        try:
            from openpyxl import load_workbook
        except Exception as exc:  # pragma: no cover
            self.skipTest(f"openpyxl unavailable: {exc}")

        with TemporaryDirectory() as temp_dir:
            exporter = Exporter(
                Settings(export_dir=Path(temp_dir)),
                FakeDb([self._finding()]),  # type: ignore[arg-type]
            )

            path = exporter.xlsx_for_run(1)
            workbook = load_workbook(path)

        self.assertEqual(workbook.sheetnames, ["Отчет", "Доказательства"])
        report_headers = [cell.value for cell in workbook["Отчет"][1]]
        evidence_headers = [cell.value for cell in workbook["Доказательства"][1]]
        self.assertNotIn("ID дела", report_headers)
        self.assertIn("SHA-256 HTML", evidence_headers)

    def test_evidence_paths_resolve_against_evidence_dir(self) -> None:
        exporter = Exporter(
            Settings(evidence_dir=Path("/var/data/evidence")),
            FakeDb([]),  # type: ignore[arg-type]
        )

        self.assertEqual(
            exporter._local_path("evidence/screenshots/example.png"),
            Path("/var/data/evidence/screenshots/example.png"),
        )

    @staticmethod
    def _finding() -> dict:
        return {
            "id": 7,
            "case_id": 3,
            "run_id": 1,
            "domain": "example.com",
            "url": "https://example.com",
            "final_url": "https://example.com/login",
            "risk_score": 82,
            "verdict": "suspicious",
            "category": "phishing",
            "case_status": "uninvestigated",
            "saved": False,
            "archived": False,
            "status_code": 200,
            "mirror_group": "",
            "title": "Example login",
            "reasons": ["Форма ввода пароля"],
            "sources": [{"url": "https://source.example/report"}],
            "screenshot_path": "evidence/screenshots/example.png",
            "html_path": "evidence/example.html",
            "html_sha256": "a" * 64,
            "created_at": "2026-07-03T00:00:00+00:00",
        }


if __name__ == "__main__":
    unittest.main()
