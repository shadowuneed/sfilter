from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

from app.config import Settings
from app.database import Database


EXPORT_COLUMNS = [
    "domain",
    "final_url",
    "risk_score",
    "category",
    "status_code",
    "mirror_group",
    "title",
    "reasons",
    "sources",
    "screenshot_file",
    "created_at",
]

TECHNICAL_EXPORT_COLUMNS = [
    "case_id",
    "run_id",
    "finding_id",
    "domain",
    "final_url",
    "verdict",
    "case_status",
    "saved",
    "archived",
    "status_code",
    "screenshot_file",
    "html_file",
    "html_sha256",
    "created_at",
]

EXPORT_HEADERS = {
    "case_id": "ID дела",
    "run_id": "ID запуска",
    "finding_id": "ID находки",
    "domain": "Домен",
    "final_url": "Адрес сайта",
    "risk_score": "Риск",
    "verdict": "Технический вердикт",
    "category": "Категория",
    "case_status": "Статус расследования",
    "saved": "Сохранено",
    "archived": "Архив",
    "status_code": "HTTP",
    "mirror_group": "Группа зеркал",
    "title": "Заголовок",
    "reasons": "Почему отмечен",
    "sources": "Источники",
    "screenshot_file": "Скриншот",
    "html_file": "Файл HTML",
    "html_sha256": "SHA-256 HTML",
    "created_at": "Дата проверки",
}

CATEGORY_LABELS = {
    "legit": "Низкий риск",
    "casino": "Казино / ставки",
    "gambling": "Казино / ставки",
    "betting": "Букмекер / ставки",
    "sports_betting_review": "Букмекер / ставки",
    "phishing": "Фишинг",
    "pyramid": "Финансовая пирамида",
    "scam": "Скам",
    "suspicious": "Подозрительный",
}

VERDICT_LABELS = {
    "suspected_fraud_or_illegal": "Высокий риск",
    "suspicious": "Нужна проверка",
    "needs_review": "Проверить вручную",
    "low_signal": "Низкий сигнал",
}


class Exporter:
    def __init__(self, settings: Settings, db: Database):
        self.settings = settings
        self.db = db

    def csv_for_run(self, run_id: int) -> Path:
        path = self.settings.export_dir / f"run_{run_id}_argus_report.csv"
        findings = self.db.list_findings(run_id=run_id, limit=10_000)
        return self._write_csv(path, findings)

    def xlsx_for_run(self, run_id: int) -> Path:
        path = self.settings.export_dir / f"run_{run_id}_argus_report.xlsx"
        findings = self.db.list_findings(run_id=run_id, limit=10_000)
        return self._write_xlsx(path, findings)

    def csv_for_cases(self, case_ids: list[int]) -> Path:
        suffix = "selected" if case_ids else "saved"
        path = self.settings.export_dir / f"cases_{suffix}_argus_report.csv"
        findings = self.db.list_findings_for_cases(case_ids)
        return self._write_csv(path, findings)

    def xlsx_for_cases(self, case_ids: list[int]) -> Path:
        suffix = "selected" if case_ids else "saved"
        path = self.settings.export_dir / f"cases_{suffix}_argus_report.xlsx"
        findings = self.db.list_findings_for_cases(case_ids)
        return self._write_xlsx(path, findings)

    def _write_csv(self, path: Path, findings: list[dict[str, Any]]) -> Path:
        with path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=[EXPORT_HEADERS[col] for col in EXPORT_COLUMNS])
            writer.writeheader()
            for finding in findings:
                row = self._row(finding)
                writer.writerow({EXPORT_HEADERS[col]: row[col] for col in EXPORT_COLUMNS})
        return path

    def _write_xlsx(self, path: Path, findings: list[dict[str, Any]]) -> Path:
        try:
            from openpyxl import Workbook
            from openpyxl.drawing.image import Image
            from openpyxl.styles import Alignment, Font, PatternFill
        except Exception:
            return self._write_csv(path.with_suffix(".csv"), findings)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Отчет"
        headers = [EXPORT_HEADERS[col] for col in EXPORT_COLUMNS]
        screenshot_col = EXPORT_COLUMNS.index("screenshot_file") + 1
        headers.insert(screenshot_col - 1, "Снимок")
        sheet.append(headers)

        header_fill = PatternFill("solid", fgColor="182432")
        header_font = Font(color="FFFFFF", bold=True)
        self._style_header(sheet, header_fill, header_font, Alignment)

        for index, finding in enumerate(findings, start=2):
            row = self._row(finding)
            values = [row[col] for col in EXPORT_COLUMNS]
            values.insert(screenshot_col - 1, "")
            sheet.append(values)
            sheet.row_dimensions[index].height = 92
            screenshot = self._local_path(row["screenshot_file"])
            if screenshot and screenshot.exists() and screenshot.suffix.lower() in {".png", ".jpg", ".jpeg"}:
                try:
                    image = Image(str(screenshot))
                    image.width = 150
                    image.height = 85
                    sheet.add_image(image, f"{self._excel_column(screenshot_col)}{index}")
                except Exception:
                    sheet.cell(index, screenshot_col).value = row["screenshot_file"]

        widths = {
            "A": 26, "B": 42, "C": 10, "D": 18, "E": 10, "F": 22,
            "G": 34, "H": 64, "I": 44, "J": 24, "K": 20,
        }
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        tech_sheet = workbook.create_sheet("Доказательства")
        tech_sheet.append([EXPORT_HEADERS[col] for col in TECHNICAL_EXPORT_COLUMNS])
        self._style_header(tech_sheet, header_fill, header_font, Alignment)
        for finding in findings:
            row = self._row(finding)
            tech_sheet.append([row[col] for col in TECHNICAL_EXPORT_COLUMNS])
        tech_widths = {
            "A": 10, "B": 10, "C": 12, "D": 26, "E": 42, "F": 22,
            "G": 20, "H": 12, "I": 10, "J": 10, "K": 34, "L": 38,
            "M": 48, "N": 20,
        }
        for column, width in tech_widths.items():
            tech_sheet.column_dimensions[column].width = width
        for row in tech_sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        workbook.save(path)
        return path

    def _row(self, finding: dict[str, Any]) -> dict[str, Any]:
        sources = finding.get("sources") or []
        reasons = finding.get("reasons") or []
        category = str(finding.get("category") or "")
        verdict = str(finding.get("verdict") or "")
        return {
            "case_id": finding.get("case_id") or "",
            "run_id": finding.get("run_id"),
            "finding_id": finding.get("id"),
            "domain": finding.get("domain"),
            "final_url": finding.get("final_url") or finding.get("url"),
            "risk_score": finding.get("risk_score"),
            "verdict": VERDICT_LABELS.get(verdict, verdict),
            "category": CATEGORY_LABELS.get(category.lower(), category),
            "case_status": finding.get("case_status") or "",
            "saved": self._yes_no(finding.get("saved")),
            "archived": self._yes_no(finding.get("archived")),
            "status_code": finding.get("status_code"),
            "mirror_group": finding.get("mirror_group") or "",
            "title": finding.get("title") or "",
            "reasons": "\n".join(str(item) for item in reasons[:8]),
            "sources": "\n".join(str(item.get("url") or item) for item in sources),
            "screenshot_file": finding.get("screenshot_path") or "",
            "html_file": finding.get("html_path") or "",
            "html_sha256": finding.get("html_sha256") or "",
            "created_at": finding.get("created_at"),
        }

    def _local_path(self, value: str | None) -> Path | None:
        if not value:
            return None
        path = Path(value)
        if path.is_absolute():
            return path
        normalized = str(value).replace("\\", "/")
        if normalized == "evidence" or normalized.startswith("evidence/"):
            relative = normalized.removeprefix("evidence").lstrip("/")
            return self.settings.evidence_dir / relative
        return Path.cwd() / path

    @staticmethod
    def _style_header(sheet: Any, fill: Any, font: Any, alignment_cls: Any) -> None:
        for cell in sheet[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = alignment_cls(wrap_text=True, vertical="top")

    @staticmethod
    def _excel_column(index: int) -> str:
        letters = ""
        while index:
            index, remainder = divmod(index - 1, 26)
            letters = chr(65 + remainder) + letters
        return letters

    @staticmethod
    def _yes_no(value: Any) -> str:
        return "да" if bool(value) else "нет"
